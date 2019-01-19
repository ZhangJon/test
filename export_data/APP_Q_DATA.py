#!usr/bin/env python
# -*- coding: utf-8 -*-

"""
@author: Jon Zhang 
@contact: zj.fly100@gmail.com
@site: 

@version: 1.1`
@license: 
@file: CC_Query_DATA.py
@time: 2018-12-29 22:33

the line began to write the explanation and demonstration of this document
"""

import xml.etree.ElementTree as ET
import sys
import os.path
import time
import base64
import logging
import json
import logging.config
import smtplib
from email.mime.text import MIMEText
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

import openpyxl
import cx_Oracle
import pymysql
import pymssql
from cryptography.hazmat.primitives import padding
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend


def SetupLogging(default_path='logging.json', default_level=logging.debug,env_key='LOG_CFG'):
    """Setup logging configuration
    """
    path = default_path
    value = os.getenv(env_key, None)
    if value:
        path = value
    if os.path.exists(path):
        with open(path, 'rt') as f:
            config = json.load(f)
        logging.config.dictConfig(config)
    else:
        logging.basicConfig(level=default_level)


def EnterTime():
    timeStruct = time.localtime()
    nowTime = time.strftime('%Y%m%d%H%M%S', timeStruct)
    time_flag = time.strftime('%Y-%m-%d %H:%M', timeStruct)
    return nowTime,time_flag


class AESCrypto(object):
    """AESCrypto."""

    def __init__(self, aes_key, aes_iv):
        """aes_key, aes_iv 可以自己定义，aes_key 32位，aes_iv 16位，如果用中文，一个中文字占三位"""
        if not isinstance(aes_key, bytes):
            aes_key = aes_key.encode()

        if not isinstance(aes_iv, bytes):
            aes_iv = aes_iv.encode()

        self.aes_key = aes_key
        self.aes_iv = aes_iv

    def encrypt(self, data, mode='cbc'):
        """encrypt."""
        func_name = '{}_encrypt'.format(mode)
        func = getattr(self, func_name)
        if not isinstance(data, bytes):
            data = data.encode()

        return func(data)

    def decrypt(self, data, mode='cbc'):
        """decrypt."""
        func_name = '{}_decrypt'.format(mode)
        func = getattr(self, func_name)

        if not isinstance(data, bytes):
            data = data.encode()

        return func(data)

    # def cfb_encrypt(self, data):
    #     """CFB encrypt."""
    #     cipher = Cipher(algorithms.AES(self.aes_key),
    #                     modes.CFB(self.aes_iv),
    #                     backend=default_backend())

    #     return cipher.encryptor().update(data)

    # def cfb_decrypt(self, data):
    #     """CFB decrypt."""
    #     cipher = Cipher(algorithms.AES(self.aes_key),
    #                     modes.CFB(self.aes_iv),
    #                     backend=default_backend())

    #     return cipher.decryptor().update(data)

    def ctr_encrypt(self, data):
        """ctr_encrypt."""
        cipher = Cipher(algorithms.AES(self.aes_key),
                        modes.CTR(self.aes_iv),
                        backend=default_backend())

        return cipher.encryptor().update(self.pkcs7_padding(data))

    def ctr_decrypt(self, data):
        """ctr_decrypt."""
        cipher = Cipher(algorithms.AES(self.aes_key),
                        modes.CTR(self.aes_iv),
                        backend=default_backend())

        uppaded_data = self.pkcs7_unpadding(cipher.decryptor().update(data))
        return uppaded_data.decode()

    def cbc_encrypt(self, data):
        """cbc_encrypt."""
        cipher = Cipher(algorithms.AES(self.aes_key),
                        modes.CBC(self.aes_iv),
                        backend=default_backend())

        return cipher.encryptor().update(self.pkcs7_padding(data))

    def cbc_decrypt(self, data):
        """cbc_decrypt."""
        cipher = Cipher(algorithms.AES(self.aes_key),
                        modes.CBC(self.aes_iv),
                        backend=default_backend())

        uppaded_data = self.pkcs7_unpadding(cipher.decryptor().update(data))
        return uppaded_data.decode()

    @staticmethod
    def pkcs7_padding(data):
        """pkcs7_padding."""
        padder = padding.PKCS7(algorithms.AES.block_size).padder()

        padded_data = padder.update(data) + padder.finalize()

        return padded_data

    @staticmethod
    def pkcs7_unpadding(padded_data):
        """pkcs7_unpadding."""
        unpadder = padding.PKCS7(algorithms.AES.block_size).unpadder()
        data = unpadder.update(padded_data)

        try:
            uppadded_data = data + unpadder.finalize()
        except ValueError:
            raise ValueError('无效的加密信息!')
        else:
            return uppadded_data


class XmlParse:
    def __init__(self, file_path):
        self.tree = None
        self.root = None
        self.xml_file_path = file_path

    def ReadXml(self):
        try:
            # print("xmlfile:", self.xml_file_path)
            self.tree = ET.parse(self.xml_file_path)
            self.root = self.tree.getroot()
        except Exception as e:
            print("parse xml faild!")
            print(e)
            sys.exit()
        else:
            pass
            # print("parse xml success!")
        finally:
            return self.tree


def ConnDb(**dict_attrib):
    """
    建立数据库连接
    """
    SetupLogging(default_path='logging.json', default_level=logging.debug, env_key='LOG_CFG')
    logger = logging.getLogger(__name__)
    try:
        database_type = dict_attrib['database_type']
        ip_adress = dict_attrib['ip_adress']
        user_name = dict_attrib['user_name']
        user_pwd = dict_attrib['user_pwd']
        logger.debug(user_pwd)
        logger.debug(type(user_pwd))
        user_pwd = base64.b64decode(user_pwd.encode('utf-8'))

        logger.debug(user_pwd)
        logger.debug(type(user_pwd))

        crypto = AESCrypto('张-密^&($#@-7/*FK-ybf码w-学f', 'm2Vd=G进yt%爱&')
        user_pwd = crypto.decrypt(user_pwd)
        db_name = dict_attrib['db_name']
        listen_port = dict_attrib['listen_port']
        if database_type == 'O':
            dsn = cx_Oracle.makedsn(ip_adress, listen_port, db_name)
            conn = cx_Oracle.connect(user_name, user_pwd, dsn)
            return conn
        elif database_type == 'M':
            conn = pymysql.connect(ip_adress, user_name, user_pwd, db_name, charset='utf8')
            return conn
        elif database_type == 'S':
            conn = pymssql.connect(ip_adress, user_name, user_pwd, db_name)
            return conn
    except Exception as e:
        logger.exception("连接数据库功能存在异常：%s" %e)


def SelectDB(db, sql):
    """
    执行sql语句，返回结果
    :param db: 数据库连接地址
    :param sql:需要执行的sql语句
    :return:返回结果
    """
    SetupLogging(default_path='logging.json', default_level=logging.debug, env_key='LOG_CFG')
    logger = logging.getLogger(__name__)
    try:
        cursor = db.cursor()
        cursor.execute(sql)
        result = cursor.fetchall()
        cursor.close()
        db.close()
        return result
    except Exception as e:
        logger.exception("数据库执行脚本功能存在异常：%s" %e)


def SendMail(workbookName,emailto,excelname):
    """
    创建带附件的邮件内容及发送详情，并连接邮件服务器发送
    :param workbookName: 创建的excel地址信息
    :return: None
    """
    SetupLogging(default_path='logging.json', default_level=logging.debug, env_key='LOG_CFG')
    logger = logging.getLogger(__name__)
    try:
        sender = 'IT-REPORT-DATA@ccb-life.com.cn'
        receiver = emailto

        subject = excelname
        smtpserver = '10.100.9.55'

        # 创建实例
        msg = MIMEMultipart()
        msg['Subject'] = Header(subject, 'utf-8')
        msg['from'] = sender
        msg['to'] = receiver

        puretext = MIMEText("""<font>[%s]查取数据请查收，有异常请第一时间联系相关人员，谢谢！</font>""" %excelname, 'html', 'utf-8')
        msg.attach(puretext)

        # 读取xlsx文件作为附件(读文件，添加，更新信息)
        xlsxpart = MIMEApplication(open('%s'%workbookName, 'rb').read())
        xlsxpart.add_header('Content-Disposition', 'attachment', filename='%s'%workbookName)
        msg.attach(xlsxpart)

        smtp = smtplib.SMTP()
        smtp.connect(smtpserver)
        smtp.sendmail(sender, receiver.split(','), msg.as_string())
        smtp.quit()
    except Exception as e:
        logger.exception("发送邮件功能存在异常：%s" %e)


def WriteNewDataToExcel(oneOfTheQueryData,oneOfTheSheetName):
    SetupLogging(default_path='logging.json', default_level=logging.debug, env_key='LOG_CFG')
    logger = logging.getLogger(__name__)
    try:
        row_len_table_name = len(oneOfTheQueryData)
        if row_len_table_name == 0:
            logger.info("[%s] 无数据结果" %oneOfTheSheetName)
            return
        column_len_table_name = len(oneOfTheQueryData[0])
        for i in range(row_len_table_name):
            oneOfTheSheetName.cell(row=i + 2, column=1).value = i+1
            for j in range(column_len_table_name):
                oneOfTheSheetName.cell(row=i + 2, column=j + 2).value = oneOfTheQueryData[i][j]
    except Exception as e:
        logger.exception("写入Excel存在异常：%s" %e)


def CreateFlag(sheet_name,time_flag):
    SetupLogging(default_path='logging.json', default_level=logging.debug, env_key='LOG_CFG')
    logger = logging.getLogger(__name__)
    try:
        sheet_values = "截止至%s业绩报告" %time_flag
        sheet_name.cell(row=23, column=1).value = sheet_values
    except Exception as e:
        logger.exception("创建日期存在异常：%s" %e)


def createExcel(nowTime,time_flag,excelname,*shuJuKuChaXunXinXi,**allSheetNameDict):
    """
    以固定模板创建一个新的excel文件
    """
    SetupLogging(default_path='logging.json', default_level=logging.debug, env_key='LOG_CFG')
    logger = logging.getLogger(__name__)
    logger.debug(allSheetNameDict)
    logger.debug(shuJuKuChaXunXinXi)
    try:
        wb = openpyxl.load_workbook('%s.xlsx'%excelname)
        for i in range(len(shuJuKuChaXunXinXi)):
            oneOfTheSheetName = wb[allSheetNameDict[str(i+1)]]
            logger.debug(shuJuKuChaXunXinXi[i])
            WriteNewDataToExcel(shuJuKuChaXunXinXi[i], oneOfTheSheetName)
        workbookName = '%s%s.xlsx' % (excelname, nowTime)
        CreateFlag(wb['报表模板参照'], time_flag)
        wb.save(workbookName)
        return workbookName
    except Exception as e:
        logger.exception("创建保存Excel存在异常：%s" %e)


def MatchData(nowTime,time_flag,dcit_sheet_sql,**dict_attrib):
    SetupLogging(default_path='logging.json', default_level=logging.debug, env_key='LOG_CFG')
    logger = logging.getLogger(__name__)
    try:
        EXCEL_NAME = dict_attrib['EXCEL_NAME']
        EMAIL_ADD = dict_attrib['EMAIL_ADD']
        logger.debug('dict_attrib：%s' %dict_attrib)
        logger.debug('EXCEL_NAME：%s' %EXCEL_NAME)
        logger.debug('EMAIL_ADDRESS：%s' %EMAIL_ADD)
        logger.debug('dcit_sheet_sql：%s' %dcit_sheet_sql)
        logger.debug('nowTime：%s' %nowTime)
        shuJuKuChaXunXinXi=[]
        allSheetNameDict={}
        for i in dcit_sheet_sql.keys():
            allSheetNameDict[i] = dcit_sheet_sql[i][0]
            sql = dcit_sheet_sql[i][1]
            # TODO 下面3行要
            dbResult = ConnDb(**dict_attrib)
            selectResult = SelectDB(dbResult, sql)
            shuJuKuChaXunXinXi.append(selectResult)

        # TODO 下面2行要
        workbookName = createExcel(nowTime, time_flag, EXCEL_NAME, *shuJuKuChaXunXinXi, **allSheetNameDict)
        # SendMail(workbookName, EMAIL_ADDRESS, EXCEL_NAME)
    except Exception as e:
        logger.error("查询数据结果处理功能存在异常：%s" %e)


def ReadMain(XML_NAME):
    SetupLogging(default_path='logging.json', default_level=logging.debug, env_key='LOG_CFG')
    logger = logging.getLogger(__name__)
    try:
        nowTime, time_flag = EnterTime()
        logger.info(XML_NAME)
        xml_file = os.path.abspath(XML_NAME)
        logger.info(xml_file)
        parse = XmlParse(xml_file)
        tree = parse.ReadXml()
        root = tree.getroot()
        logger.debug(root)
        captionList = root.findall("EXCEL")  # 在当前指定目录下遍历
        logger.debug(len(captionList))
        for caption in captionList:
            dict_attrib = caption.attrib
            logger.debug("%s----%s----%s" %(caption.tag, caption.attrib, caption.text))
            child_list = caption.findall("SQL")
            dcit_sheet_sql = {}
            for i in range(len(child_list)):
                dcit_sheet_sql[str(i+1)] = [child_list[i].attrib['sheet_name'], child_list[i].text]
            MatchData(nowTime,time_flag,dcit_sheet_sql,**dict_attrib)
    except Exception as e:
        logger.error("调用函数功能存在异常：%s" %e)


if __name__ == "__main__":
    ReadMain('EXP_DATA_CONF.xml')
