#!usr/bin/env python
# -*- coding: utf-8 -*-

"""
@author: Jon Zhang 
@contact: zj.fly100@gmail.com
@site: 

@version: 1.1`
@license: 
@file: CC_Query_DATA.py
@time: 2018-12-29 22:33

在已有模板更新数据，并把最新的数据发送到指定的对象
"""

import xml.etree.ElementTree as ET
import sys
import time
import datetime
import os.path
import time
import base64
import logging
import json
import logging.config
import smtplib
from email.mime.text import MIMEText
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

import openpyxl
import cx_Oracle
import pymysql
import pymssql
from cryptography.hazmat.primitives import padding
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend


def SetupLogging(default_path='logging.json', default_level=logging.debug,env_key='LOG_CFG'):
    """Setup logging configuration
    """
    path = default_path
    value = os.getenv(env_key, None)
    if value:
        path = value
    if os.path.exists(path):
        with open(path, 'rt') as f:
            config = json.load(f)
        logging.config.dictConfig(config)
    else:
        logging.basicConfig(level=default_level)


def EnterTime():
    timeStruct = time.localtime()
    nowTime = time.strftime('%Y%m%d%H%M%S', timeStruct)
    time_flag = time.strftime('%Y-%m-%d %H:%M', timeStruct)
    return nowTime,time_flag


def timeer(func):
    def wrapper(*args, **kwargs):
        SetupLogging(default_path='logging.json', default_level=logging.debug, env_key='LOG_CFG')
        logger = logging.getLogger(__name__)
        a = datetime.datetime.now()
        func(*args, **kwargs)
        b = datetime.datetime.now()
        logger.debug("使用时间：%s" % (b - a))
    return wrapper


class AESCrypto(object):
    """AESCrypto."""

    def __init__(self, aes_key, aes_iv):
        """aes_key, aes_iv 可以自己定义，aes_key 32位，aes_iv 16位，如果用中文，一个中文字占三位"""
        if not isinstance(aes_key, bytes):
            aes_key = aes_key.encode()

        if not isinstance(aes_iv, bytes):
            aes_iv = aes_iv.encode()

        self.aes_key = aes_key
        self.aes_iv = aes_iv

    def encrypt(self, data, mode='cbc'):
        """encrypt."""
        func_name = '{}_encrypt'.format(mode)
        func = getattr(self, func_name)
        if not isinstance(data, bytes):
            data = data.encode()

        return func(data)

    def decrypt(self, data, mode='cbc'):
        """decrypt."""
        func_name = '{}_decrypt'.format(mode)
        func = getattr(self, func_name)

        if not isinstance(data, bytes):
            data = data.encode()

        return func(data)

    # def cfb_encrypt(self, data):
    #     """CFB encrypt."""
    #     cipher = Cipher(algorithms.AES(self.aes_key),
    #                     modes.CFB(self.aes_iv),
    #                     backend=default_backend())

    #     return cipher.encryptor().update(data)

    # def cfb_decrypt(self, data):
    #     """CFB decrypt."""
    #     cipher = Cipher(algorithms.AES(self.aes_key),
    #                     modes.CFB(self.aes_iv),
    #                     backend=default_backend())

    #     return cipher.decryptor().update(data)

    def ctr_encrypt(self, data):
        """ctr_encrypt."""
        cipher = Cipher(algorithms.AES(self.aes_key),
                        modes.CTR(self.aes_iv),
                        backend=default_backend())

        return cipher.encryptor().update(self.pkcs7_padding(data))

    def ctr_decrypt(self, data):
        """ctr_decrypt."""
        cipher = Cipher(algorithms.AES(self.aes_key),
                        modes.CTR(self.aes_iv),
                        backend=default_backend())

        uppaded_data = self.pkcs7_unpadding(cipher.decryptor().update(data))
        return uppaded_data.decode()

    def cbc_encrypt(self, data):
        """cbc_encrypt."""
        cipher = Cipher(algorithms.AES(self.aes_key),
                        modes.CBC(self.aes_iv),
                        backend=default_backend())

        return cipher.encryptor().update(self.pkcs7_padding(data))

    def cbc_decrypt(self, data):
        """cbc_decrypt."""
        cipher = Cipher(algorithms.AES(self.aes_key),
                        modes.CBC(self.aes_iv),
                        backend=default_backend())

        uppaded_data = self.pkcs7_unpadding(cipher.decryptor().update(data))
        return uppaded_data.decode()

    @staticmethod
    def pkcs7_padding(data):
        """pkcs7_padding."""
        padder = padding.PKCS7(algorithms.AES.block_size).padder()

        padded_data = padder.update(data) + padder.finalize()

        return padded_data

    @staticmethod
    def pkcs7_unpadding(padded_data):
        """pkcs7_unpadding."""
        unpadder = padding.PKCS7(algorithms.AES.block_size).unpadder()
        data = unpadder.update(padded_data)

        try:
            uppadded_data = data + unpadder.finalize()
        except ValueError:
            raise ValueError('无效的加密信息!')
        else:
            return uppadded_data


class XmlParse:
    def __init__(self, file_name):
        SetupLogging(default_path='logging.json', default_level=logging.debug, env_key='LOG_CFG')
        self.logger = logging.getLogger(__name__)
        self.tree = None
        self.root = None
        self.file_path_name = os.path.abspath(file_name)
        self.dict_result = {}

    def ReadXmlTree(self):
        try:
            self.tree = ET.parse(self.file_path_name)
            self.root = self.tree.getroot()
        except Exception as e:
            self.logger.error("parse xml faild!\n %s" %e)
            sys.exit()
        else:
            self.logger.info("parse xml success!")
        finally:
            return self.tree, self.root

    def CreateDictResults(self):
        self.ReadXmlTree()
        try:
            captionList = self.root.findall("EXCEL")  # 在当前指定目录下遍历
            self.logger.debug(len(captionList))
            for caption in captionList:
                dcit_sheet_name_sql = {}
                self.dict_result = caption.attrib
                self.logger.debug("%s----%s----%s" % (caption.tag, caption.attrib, caption.text))
                child_list = caption.findall("SQL")
                for i in range(len(child_list)):
                    dcit_sheet_name_sql[child_list[i].attrib['sheet_name']] = child_list[i].text
                self.dict_result['dcit_sheet_name_sql'] = dcit_sheet_name_sql
        except Exception as e:
            self.logger.error("调用ReadMain函数功能存在异常：%s" % e)
        finally:
            return self.dict_result


class DatabaseOperation:
    def __init__(self,*args,**kwargs):
        self.conn = None
        self.args = args
        self.kwargs = kwargs
        SetupLogging(default_path='logging.json', default_level=logging.debug, env_key='LOG_CFG')
        self.logger = logging.getLogger(__name__)

    def ConnDatabase(self):
        database_type = self.kwargs['database_type']
        ip_adress = self.kwargs['ip_adress']
        user_name = self.kwargs['user_name']
        user_pwd = self.kwargs['user_pwd']
        db_name = self.kwargs['db_name']
        listen_port = self.kwargs['listen_port']
        user_pwd = base64.b64decode(user_pwd.encode('utf-8'))
        crypto = AESCrypto('张-密^&($#@-7/*FK-ybf码w-学f', 'm2Vd=G进yt%爱&')
        user_pwd = crypto.decrypt(user_pwd)
        if database_type == 'O':
            dsn = cx_Oracle.makedsn(ip_adress, listen_port, db_name)
            self.conn = cx_Oracle.connect(user_name, user_pwd, dsn)
            return self.conn
        elif database_type == 'S':
            self.conn = pymssql.connect(ip_adress, user_name, user_pwd, db_name)
            return self.conn
        elif database_type == 'M':
            self.conn = pymysql.connect(ip_adress, user_name, user_pwd, db_name, charset='utf8')
            return self.conn

    def ExecuteDatabaseOperation(self):
        sql = self.kwargs['sql']
        self.logger.debug(sql)
        cursor = self.ConnDatabase().cursor()
        cursor.execute(sql)
        yield cursor
        # results = cursor.fetchall()
        # self.logger.debug(results)
        cursor.close()
        self.conn.close()
        # return results


class ExcelOperation:
    def __init__(self,**kwargs):
        SetupLogging(default_path='logging.json', default_level=logging.debug, env_key='LOG_CFG')
        self.logger = logging.getLogger(__name__)
        self.wb = None
        self.kwargs = kwargs

    def CreateExcel(self):
        self.logger.debug(self.kwargs)
        excelname = self.kwargs['EXCEL_NAME']
        nowTime = self.kwargs['nowTime']
        time_flag = self.kwargs['time_flag']
        dcit_sheet_name_sql = self.kwargs['dcit_sheet_name_sql']
        new_workbook_name = '%s%s.xlsx' % (excelname, nowTime)
        self.wb = openpyxl.load_workbook('%s.xlsx' % excelname)
        for oneOfTheSheetName in dcit_sheet_name_sql.keys():
            self.WriteNewDataToExcel(dcit_sheet_name_sql[oneOfTheSheetName], self.wb[oneOfTheSheetName])
            # self.wb.save(new_workbook_name)

        self.CreateFlag(self.wb['报表模板参照'], time_flag)
        self.wb.save(new_workbook_name)
        self.wb.close()
        return new_workbook_name

    def WriteNewDataToExcel(self,oneOfTheQueryData, oneOfTheSheetName):
        self.logger.debug(oneOfTheQueryData)
        row_num = 0
        for cursor in oneOfTheQueryData:
            self.logger.debug(cursor)
            the_first_row_num = cursor.fetchone()
            self.logger.debug(the_first_row_num)
            if the_first_row_num == None:
                self.logger.info("[%s] 无数据结果" % oneOfTheSheetName)
                continue
            for j in range(len(the_first_row_num)):
                oneOfTheSheetName.cell(row=row_num + 2, column=j + 1).value = the_first_row_num[j]
            for i in cursor:
                row_len_table_name = len(i)
                for j in range(row_len_table_name):
                    oneOfTheSheetName.cell(row=row_num + 3, column=j + 1).value = i[j]
                row_num += 1

    def CreateFlag(self,sheet_name, time_flag):
        try:
            sheet_values = "截止至%s业绩报告" % time_flag
            sheet_name.cell(row=23, column=1).value = sheet_values
        except Exception as e:
            self.logger.error("创建日期存在异常：%s" % e)


def SendMail(**kwargs):
    """
    创建带附件的邮件内容及发送详情，并连接邮件服务器发送
    """
    SetupLogging(default_path='logging.json', default_level=logging.debug, env_key='LOG_CFG')
    logger = logging.getLogger(__name__)
    EXCEL_NAME = kwargs['EXCEL_NAME']
    EMAIL_ADD = kwargs['EMAIL_ADD']
    workbookName = kwargs['workbookName']
    try:
        sender = 'IT-REPORT-DATA@ccb-life.com.cn'
        receiver = EMAIL_ADD

        subject = EXCEL_NAME
        smtpserver = '邮件服务器地址'

        # 创建实例
        msg = MIMEMultipart()
        msg['Subject'] = Header(subject, 'utf-8')
        msg['from'] = sender
        msg['to'] = receiver

        puretext = MIMEText("""<font>[%s]查取数据请查收，有异常请第一时间联系相关人员，谢谢！</font>""" %EXCEL_NAME, 'html', 'utf-8')
        msg.attach(puretext)

        # 读取xlsx文件作为附件(读文件，添加，更新信息)
        xlsxpart = MIMEApplication(open('%s'%workbookName, 'rb').read())
        xlsxpart.add_header('Content-Disposition', 'attachment', filename='%s'%workbookName)
        msg.attach(xlsxpart)

        smtp = smtplib.SMTP()
        smtp.connect(smtpserver)
        smtp.sendmail(sender, receiver.split(','), msg.as_string())
        smtp.quit()
    except Exception as e:
        logger.error("发送邮件功能存在异常：%s" %e)


@timeer
def Main(XML_NAME):
    SetupLogging(default_path='logging.json', default_level=logging.debug, env_key='LOG_CFG')
    logger = logging.getLogger(__name__)
    nowTime, time_flag = EnterTime()
    dict_results = XmlParse(XML_NAME).CreateDictResults()
    dict_results['nowTime'] = nowTime
    dict_results['time_flag'] = time_flag
    dcit_sheet_name_sql = dict_results['dcit_sheet_name_sql']
    logger.debug('dict_attrib：%s' % dict_results)
    logger.debug('dcit_sheet_sql：%s' % dict_results)
    logger.debug('nowTime：%s' % nowTime)

    for i in dcit_sheet_name_sql.keys():
        dict_results['sql'] = dcit_sheet_name_sql[i]
        DatabaseResultsHandle = DatabaseOperation(**dict_results)
        DatabaseResults = DatabaseResultsHandle.ExecuteDatabaseOperation()
        dcit_sheet_name_sql[i] = DatabaseResults
        logger.debug('DatabaseResults:%s' %DatabaseResults)

    # TODO 下面3行要
    dict_results['workbookName'] = ExcelOperation(**dict_results).CreateExcel()
    # SendMail(**dict_results)

if __name__ == "__main__":
    Main('EXP_DATA_CONF.xml')
